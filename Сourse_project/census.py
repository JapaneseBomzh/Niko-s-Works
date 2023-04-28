import openpyxl
from adress_list import def_address,def_Oblast
from question_list import question_list
question,error=question_list()

#Ім'я
print(question[0])
name = input()

#Прізвище
print(question[1])
surname = input()
 
#По батькові
print(question[2])
fathers_name_q = input()
while fathers_name_q not in ('1', '2'):
    print(error[0])
    print(question[2])
    fathers_name_q = input()
if fathers_name_q == "1":
    print(question[3])
    fathers_name = input()
else:
    fathers_name ="-"

#Номер телефону
print(question[4])
phone = input()
while len(phone) != 10 or not phone.isdigit():
    print(error[1])
    print(question[4])
    phone = input()

#Область,Район,Адреса
print(question[5])
print(question[6])
oblasti=def_Oblast(0)
while True:
    for i, oblast in enumerate(oblasti):
        print(f"{i+1} - {oblast}")
    try:
        num = int(input())
        if num < 1 or num > len(oblasti):
            print(error[2], len(oblasti))
            continue
        oblast = oblasti[num-1]
        break
    except ValueError:
        print(error[2], len(oblasti))
        continue
print(question[7])
rayoni=def_address(oblast)
while True:
    try:
        for i, rayon in enumerate(rayoni):
            print(f"{i+1} - {rayon}")
        rayon = rayoni[int(input()) - 1]
        break
    except IndexError:
            print(error[3])
    except ValueError:
            print(error[4])
print(question[8])
home_adress=input()

#2 - Стать
print(question[9])
gender = input()
while gender not in ('1', '2', '3', '4', '5'):
    print(error[5])
    print(question[9])
    gender = input()

#4 - Дата Народження
print(question[10])
birthdate = input()
while True:
    try:
        day, month, year = map(int, birthdate.split('/'))
        if len(birthdate) != 10 or birthdate[2] != '/' or birthdate[5] != '/' or not (1 <= month <= 12) or not (1 <= day <= 31):
            raise ValueError
        break
    except ValueError:
        print(error[6])
        print(question[10])
        birthdate = input()

#4 - Скільки виповнилось років
print(question[11])
age = int(input())

#5 - Місце народження
print(question[12])
birthplaces = def_Oblast(1)
while True:
    for i, birthplace in enumerate(birthplaces):
        print(f"{i+1} - {birthplace}")
    try:
        print(question[6])
        num = int(input())
        if num < 1 or num > len(birthplaces):
            print(error[2], len(birthplaces))
            continue
        birthplace = birthplaces[num-1]
        break
    except ValueError:
        print(error[2], len(birthplaces))
    continue
if birthplace == "Інша держава":
    print(question[13])
    birthplace = input()
 
#6 - Етнічне походження
print(question[14])
ethnicity = input()

#7 - Рідна мова
print(question[15])
native_language = input()
if native_language.lower() != 'українська': #7 - Володіння українською мовою
    print(question[16])
    ukrainian = input()
else:
    ukrainian = "1"
print(question[17])
other_language = input()#7 - Інша мова

#8 - Громадянство
print(question[18])
сitizenship = input()
while сitizenship not in ('1', '2', '3'): 
    print(error[0])
    print(question[18])
    сitizenship  = input()
if сitizenship == "3":
    print(question[19])
    сitizenship=input()

#9 - Сімейний стан
if age >= 15:    
    marital_statuses = ["1", "2", "3", "4", "5", "6"]
    print(question[20])
    marital_status = input()
    while marital_status not in marital_statuses:
        print(error[7])
        print(question[20])
        marital_status = input()
else:marital_status ="-"

#10 - Освіта
if age >= 6:
    print(question[21])
    education = input()
    while education not in ["1", "2", "3", "4", "5", "6", "7", "8", "9"]:
        print(error[8])
        print(question[21])
        education = input()
    while True:
        try:
            print(question[22])
            education_info = input()
            if education_info == "-":
                school = "-";classes = "-";education_year = "-"
                break
            else:
                school, classes, education_year = education_info.split(", ")
                break
        except ValueError:
            print(error[9])
else: education = "-"; education_info = "-"

#11 - Чи закінчили п-т навч. заклад? 
if age >= 15:
    print(question[23])    
    vocational_school = input()
    while vocational_school not in ["1", "2"]:
        print(error[10])
        print(question[23])
        vocational_school = input()    
else: vocational_school ="-"

#12 - Тип навчального закладу
if age >= 6: #???
    print(question[24]) 
    education_type = input()
    while education_type not in ["1", "2", "3", "4", "5", "6", "7"]:
         print(error[11])
         print(question[24]) 
         education_type = input()
else: education_type ="-"

#13 - Джерела засобів існування
set_chosen_sources = set()
while True:
    print(question[25]) 
    chosen_source = input()
    if chosen_source.isdigit() and chosen_source in ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15'):
        set_chosen_sources.add(chosen_source)
    else:
        print(error[12])
    chose = input(f"Обрані джерела засобів існування: {set_chosen_sources}\nБажаєте додати ще? (1 - так, 2 - ні): ")
    while chose not in ('1', '2'):
        print(error[13])
        chose = input(f"Обрані джерела засобів існування: {set_chosen_sources}\nБажаєте додати ще? (1 - так, 2 - ні): ")
    if chose == '2':
        break

#13 - Номер основного джерела засобів існування
print(question[26]) 
main_source = input()
while main_source not in set_chosen_sources:
    print(error[13]) 
    print(question[26]) 
    main_source = input()

#Збереження даних
wb = openpyxl.load_workbook('census.xlsx')
ws = wb.active

row = ws.max_row + 1
ws.cell(row=row, column=1, value=row-1)
ws.cell(row=row, column=2, value=name)
ws.cell(row=row, column=3, value=surname)
ws.cell(row=row, column=4, value=fathers_name)
ws.cell(row=row, column=5, value=phone)
ws.cell(row=row, column=6, value=oblast)
ws.cell(row=row, column=7, value=rayon)
ws.cell(row=row, column=8, value=home_adress)
ws.cell(row=row, column=9, value=gender)
ws.cell(row=row, column=10, value=birthdate)
ws.cell(row=row, column=11, value=birthplace)
ws.cell(row=row, column=12, value=ethnicity)
ws.cell(row=row, column=13, value=native_language)
ws.cell(row=row, column=14, value=ukrainian)
ws.cell(row=row, column=15, value=other_language)
ws.cell(row=row, column=16, value=сitizenship)
ws.cell(row=row, column=17, value=marital_status)
ws.cell(row=row, column=18, value=education)
ws.cell(row=row, column=19, value=education_info)
ws.cell(row=row, column=20, value=vocational_school)
ws.cell(row=row, column=21, value=education_type)
ws.cell(row=row, column=22, value=', '.join(set_chosen_sources))
ws.cell(row=row, column=23, value=main_source)
wb.save('census.xlsx')
